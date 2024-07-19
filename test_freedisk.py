import os
import win32file
import openpyxl
import datetime
from openpyxl.styles import Alignment, Font
class Log :
    def __init__(self,a,b,c):
        self.host = a
        self.c = b
        self.d = c
        
apache_logs = [
	Log('192.168.160.61', "D", "E"),
	Log('192.168.160.62', "H", "I"),	
	Log('192.168.160.63', "J", "K"),
	Log('192.168.160.64', "M", "N"),
	Log('192.168.160.65', "P", "Q"),
	Log('192.168.160.66', "R", "S"),
	Log('192.168.160.67', "T", "U"),
	Log('192.168.160.68', "X", "Y"),
	Log('192.168.160.73',"Z", "AA"),
	Log('192.168.160.74', "AB", "AC"),
	Log('192.168.160.76', "AD", "AE"),
    Log('192.168.160.77', "AF", "AG"),
    Log('192.168.160.81', "AH", "AI"),
    Log('192.168.160.83',  "AL", "AM"),
    Log('192.168.160.84', "AN", "AO"),
    Log('192.168.160.85', "AP", "AQ"),
    Log('192.168.160.86', "AR", "AS"),
    Log('192.168.160.87', "AU", "AV"),
    Log('192.168.160.47', "AX", None),
    Log('192.168.160.49', "AY", None),
    Log('192.168.160.50', "AZ", None),
    # Log('192.168.160.91', "D", "E"),
    # Log('192.168.160.92', "H", "I"),
    # Log('192.168.160.93', "J", "K"),
    # Log('192.168.160.94', "M", "N"),
    
]
# Lấy dung lượng còn trống của ổ đĩa qua UNC path
def get_disk_space_unc(path):
    try:
        sectors_per_cluster, bytes_per_sector, num_free_clusters, total_num_clusters = win32file.GetDiskFreeSpace(path)
        free_bytes = num_free_clusters * sectors_per_cluster * bytes_per_sector
        free_space_gb = free_bytes / (1024 ** 3)  # Đổi từ byte sang GB
        return round(free_space_gb, 1)  # Làm tròn đến 1 chữ số thập phân
    except Exception as e:
        print(f"Error accessing {path}: {e}")
        return None  # Nếu không thể truy cập
line_path = r"C:\Users\DSP189\Desktop\line2.txt"

def increase_line(line):
    return line + 1

def save_line(line):
    with open(line_path, "w") as file:
        file.write(str(line))
def read_value():
    try:
        with open(line_path, "r") as file:
            return int(file.read())
    except FileNotFoundError:
        return 0
line = read_value() # row Excel -2
# Ghi dữ liệu vào file Excel
def write_to_excel(disk_info, excel_path):
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    # Kiểm tra nếu sheet đã tồn tại
    if 'disk容量' in workbook.sheetnames:
        sheet = workbook['disk容量']
    else:
        sheet = workbook.create_sheet(title='disk容量')
    
    # Ghi dữ liệu
    
    if "C:" in disk_info and disk_info["C:"] is not None:
        cell = sheet[f'{log.c}{line}'] 
        cell.value = disk_info["C:"]
        if disk_info["C:"] < 20:
            cell.font = Font(color="FF0000")
    
    if "D:" in disk_info and disk_info["D:"] is not None:
        cell = sheet[f'{log.d}{line}'] 
        cell.value = disk_info["D:"]
        if disk_info["D:"] < 20:
            cell.font = Font(color="FF0000")
    today = datetime.datetime.today().strftime('%Y/%m/%d')
    cell =sheet[f'C{line}']
    cell.value = today
    cell.alignment = Alignment(horizontal='center')
    
    # Lưu file Excel
    workbook.save(excel_path)
# for log in apache_logs:
# Địa chỉ UNC của máy tính khác và ổ đĩa mạng
for log in apache_logs:
    remote_computer = f'\\\\{log.host}'



# Danh sách các ổ đĩa cần kiểm tra trên máy tính từ xa
    remote_drives = {
        "C:": os.path.join(remote_computer, "C$"),
        "D:": os.path.join(remote_computer, "D$")
    }

# Lấy thông tin dung lượng các ổ đĩa qua UNC path
    disk_info = {drive: get_disk_space_unc(path) for drive, path in remote_drives.items()}

# Đường dẫn tới file Excel
    excel_path = r'\\192.168.160.6\usbdisk3\システム運用\●ハード環境\121029_サーバ再起動手順書●\●新システム_サーバ停止＆起動手順_v3.1.xlsx'

# Ghi thông tin vào file Excel
    write_to_excel(disk_info, excel_path)
line = increase_line(line)# increase_rowを呼び込み、row　1個増える
     #新値の rowを　row.txtに保存
save_line(line)

