import psycopg2
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font
import openpyxl as xl

# データベースへの接続用設定
users = 'unionplate'			# ユーザID 
host = '192.168.160.83'			# 接続先IPアドレス
dbnames = 'union_hanbai'		# DB名
passwords = 'etalpnoinu'		# パスワード

def get_date_range(year, month):
    # Tính toán ngày bắt đầu (1st day of the month)
    start_date = datetime.date(year, month, 1)
    
    # Tính toán ngày kết thúc (last day of the month)
    next_month = start_date.replace(month=start_date.month+1, day=1)
    end_date = next_month - datetime.timedelta(days=1)
    
    return start_date, end_date

previous_month = datetime.date.today().month-1 
previous_year = datetime.date.today().year 
if previous_month <= 0: 
    previous_month +=12
    previous_year = previous_year - 1
print( previous_month,previous_year)
# Sử dụng hàm để lấy ngày bắt đầu và ngày kết thúc
start_date, end_date = get_date_range(previous_year, previous_month)
start_date = start_date.strftime("%Y/%m/%d")
end_date = end_date.strftime("%Y/%m/%d")
excel_date = datetime.datetime.today().replace(day=1) - datetime.timedelta(days=1)
excel_date_str = excel_date.strftime("%Y%m")

def sql_count_juchu(users, host, dbnames, passwords, staff_department_mapping, tables, start_date, end_date):
    # DB kết nối
    conn = psycopg2.connect(f"user={users} host={host} dbname={dbnames} password={passwords}")
    
    all_results = []
    for table, new_table_name in tables.items():
        # Tạo phần truy vấn SQL dựa trên tập hợp staff_names
        sql_query = f"SELECT "

        for staff_name, eigyou_bu in staff_department_mapping.items():
            sql_query += f"count(CASE WHEN create_staff in ({staff_name}) THEN 1 ELSE NULL END) as {eigyou_bu}, "

        sql_query = sql_query[:-2]  # Loại bỏ dấu phẩy cuối cùng
        sql_query += f" FROM {table} WHERE create_date BETWEEN '{start_date}' AND '{end_date}'"

        # Thực hiện truy vấn SQL
        cur = conn.cursor()
        cur.execute(sql_query)
        results = cur.fetchall()
        # Lưu kết quả vào danh sách
        all_results.append(results)
        
        # In kết quả
        print(f"Kết quả cho bảng {table}:")
        for row in results:
            for eigyou_bu, count in zip(staff_department_mapping.values(), row):
                print(f"  {eigyou_bu}: {count}")

        # Đóng con trỏ
        cur.close()

    # Đóng kết nối
    conn.close()
    return all_results

def write_results_to_excel(results, staff_department_mapping, tables, excel_path):
    # Tạo một đối tượng Workbook mới
    wb = Workbook()

    # Lấy sheet active (sheet đang hoạt động) mặc định
    ws = wb.active
    ws.title = "各部明細数"
    
    # Ghi tên bộ phận vào hàng đầu tiên
    for idx, eigyou_bu in enumerate(staff_department_mapping.values(), start=1):
        ws.cell(row=1, column=idx+1, value=eigyou_bu)

    # Ghi tên bảng vào cột A
    for idx, (table, new_table_name) in enumerate(tables.items(), start=2):
        ws.cell(row=idx, column=1, value=new_table_name)

    # Ghi kết quả vào các ô tương ứng
    for idx_row, row in enumerate(results, start=2):
        for idx_col, counts in enumerate(row, start=2):
            for count in counts:  # Duyệt qua các giá trị trong tuple
                ws.cell(row=idx_row, column=idx_col, value=count)
                idx_col += 1

    # Lưu file Excel
    wb.save(excel_path)
    print(f"Đã ghi kết quả vào file Excel thành công tại '{excel_path}'.")

def sql_juchu(users, host, dbnames, passwords, staff_department_mapping, tables, start_date, end_date, excel_path):
    # DB kết nối
    conn = psycopg2.connect(f"user={users} host={host} dbname={dbnames} password={passwords}")
    
    # Mở file Excel hiện có để ghi vào
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
        for table, new_table_name in tables.items():
            for staff_name, eigyou_bu in staff_department_mapping.items():
                sql_query = f"SELECT * FROM {table} WHERE create_staff in ({staff_name}) \
                             AND create_date BETWEEN '{start_date}' AND '{end_date}' \
                             ORDER BY create_date, create_time, create_staff, update_date, update_time, update_staff"
                data = pd.read_sql_query(sql_query, conn)
                
                # Ghi dữ liệu vào sheet mới với tên là tên bảng
                sheet_name = f"{new_table_name}_{eigyou_bu}"
                data.to_excel(writer, sheet_name=sheet_name, index=False)
                
    # Đóng kết nối
    conn.close()

# Tạo danh sách các bảng cần thực hiện truy vấn
tables = {
    "offc_trn_juchu_details": "受注明細",
    "offc_trn_mitsumori_details": "見積明細",
    "offc_trn_hachu": "発注明細",
    "offc_trn_juchu_details_check": "チェック明細"
}

# Tạo ánh xạ giữa tên nhân viên và tên bộ phận
staff_department_mapping = {}
file_path = r"C:\Users\DSP189\Desktop\test\test.txt"
# Mở tệp văn bản để đọc
with open(file_path, 'r') as file:
    for line in file:
        # Tìm vị trí của dấu ngoặc kép và dấu hai chấm trong dòng
        quote_index = line.find('"')
        colon_index = line.find(':')
        
        # Tách tên nhân viên và bộ phận từ dòng
        employee_name = line[quote_index+1:colon_index].strip()
        department = line[colon_index+1:].strip()
        
        # Lưu thông tin vào từ điển
        staff_department_mapping[employee_name] = department

# Thực hiện các truy vấn và ghi kết quả vào file Excel
results = sql_count_juchu(users, host, dbnames, passwords, staff_department_mapping, tables, start_date, end_date)
excel_date_str = pd.to_datetime(end_date).strftime('%Y%m%d')
excel_path = f"C:\\Users\\DSP189\Desktop\\test\\営業本部明細入力数_{excel_date_str}.xlsx"
write_results_to_excel(results, staff_department_mapping, tables, excel_path)
sql_juchu(users, host, dbnames, passwords, staff_department_mapping, tables, start_date, end_date, excel_path)
wb1 = xl.load_workbook(filename=excel_path)


# set font
font = Font(name='游ゴシック')

for ws1 in wb1.worksheets:
    for row in ws1:
        for cell in row:
            ws1[cell.coordinate].font = font

# save xlsx file
wb1.save(excel_path)


