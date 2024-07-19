import psycopg2
import openpyxl

class FactoryInfo:
	def __init__(self, coporation_code, office_factory_code, office_factory_name, office_factory_ryaku, ip_address, sheet_name, zaishitsu_top_cell, zaishitsu_last_cell, edit_mode):
		self.coporation_code = coporation_code
		self.factory_code = office_factory_code
		self.name = office_factory_name
		self.ryaku = office_factory_ryaku
		self.ip = ip_address
		self.sheet_name = sheet_name
		self.zaishitsu_top_cell = zaishitsu_top_cell
		self.zaishitsu_last_cell = zaishitsu_last_cell
		self.edit_mode = edit_mode

# DB接続 → SQL実行 → DB切断
def sql_run(users, host, dbnames, passwords, factory_code):
	# DB接続
	conn = psycopg2.connect(" user=" + users +" host=" + host +" dbname=" + dbnames +" password=" + passwords)

	# sqlの実行
	cur = conn.cursor()
	cur.execute(f"\
				select zaishitsu_code, sum(juryo) \
				from offc_trn_hazai_shuno_tana \
				where hazai_office_factory_code='{factory_code}' \
				group by zaishitsu_code \
				order by zaishitsu_code"
				)
	results = cur.fetchall()

	# DB切断
	cur.close()
	conn.close()

	return results
row_path = r"C:\Users\DSP189\Desktop\row.txt"
def increase_row(row):
    return row + 1

def save_row(row):
    with open(row_path, "w") as file:
        file.write(str(row))

def read_value():
    try:
        with open(row_path, "r") as file:
            return int(file.read())
    except FileNotFoundError:
        return 0

# テキストでrowの値読み込む
row = read_value()


# increase_rowを呼び込み、row　1個増える
row = increase_row(row)


# 新値の rowを　row.txtに保存
save_row(row)
# 工場の情報
factorys = [
	# FactoryInfo(コーポレートコード, 工場コード, 工場名, 工場名略, IPアドレス, シート名, 材質コード先頭セル, 材質コード最終セル, 編集モード),
	# 編集モード 0:ExcelにあってDBにない材質に0を入力　1:ExcelにあってDBにないものは何もしない
	FactoryInfo(10, 100, "㈱ユニオンプレート本社工場", "本社", "192.168.162.51", "本社全鋼種", 5, 144, 0),
	FactoryInfo(10, 101, "ユニオン本社ＦＣ工場", "FC工場", "192.168.162.61", "本社全鋼種", 5, 144, 1),
	FactoryInfo(10, 520, "ユニオン尾道アルミ工場", "尾道アルミ", "192.168.17.61", "尾道アルミ", 20, 24, 0),
	FactoryInfo(10, 570, "ユニオン尾道工場", "尾道", "192.168.17.51", "尾道全鋼種", 5, 49, 0),
	FactoryInfo(10, 580, "ユニオン厚木工場", "厚木", "192.168.18.51", "厚木全鋼種", 5, 30, 0),
	FactoryInfo(10, 650, "ユニオン小牧工場", "小牧", "192.168.25.51", "小牧全鋼種", 5, 30, 0),
	FactoryInfo(10, 670, "ユニオン上田工場", "上田", "192.168.161.51", "上田全鋼種", 5, 51, 0),
	FactoryInfo(10, 680, "ユニオン千曲工場", "千曲", "192.168.28.51", "", 0, 0, 0),
	FactoryInfo(50, 500, "ＩＮＳ加工工場", "INS", "192.168.10.51", "INS", 20, 24, 0),
	FactoryInfo(51, 510, "㈲藤精工", "藤精工", "192.168.11.51", "", 0, 0, 0),
	FactoryInfo(53, 530, "㈲サンテック", "ｻﾝﾃｯｸ", "192.168.13.51", "サンテック", 20, 31, 0),
	FactoryInfo(54, 540, "㈲エムケイ精工", "MK", "192.168.14.51", "", 0, 0, 0),
	FactoryInfo(55, 550, "㈱渡辺製作所", "渡辺", "192.168.15.51", "", 0, 0, 0),
	FactoryInfo(59, 590, "㈲エムテーエス", "MTS", "192.168.19.51", "", 0, 0, 0),
	FactoryInfo(60, 600, "ＡＵＫプレート㈱", "AUK", "192.168.20.51", "AUK", 20, 36, 0),
	FactoryInfo(60, 601, "ＡＵＫプレート㈱アルミ工場", "ＡＵＫアルミ", "192.168.20.61", "AUKアルミ", 20, 23, 0),
	FactoryInfo(63, 630, "㈱メカニックメタル", "MM", "192.168.23.51", "", 23, 24, 0),
	FactoryInfo(64, 640, "㈲鈴木鋼管工業", "鈴木", "192.168.24.51", "", 0, 0, 0)
]

# データベースへの接続用設定
users = 'unionplate'			# ユーザID 
host = '192.168.160.83'			# 接続先IPアドレス
dbnames = 'union_hanbai'		# DB名
passwords = 'etalpnoinu'		# パスワード

# Excelファイルの場所
excel_file = r'\\192.168.160.6\\usbdisk3\\システム運用\\■新システム\\端材管理システム関連\\工場別月末端材重量.xlsx'

# Excelファイルのオープン
# ブックを変数に格納
print("エクセルファイルをオープンします")
try:
	write_wb = openpyxl.load_workbook(excel_file)
except(FileNotFoundError):
	print("書き込むファイルがありません。")
	exit()

for factory in factorys:
	# シートを変数に名前で格納
	if factory.sheet_name:
		write_ws = write_wb[factory.sheet_name]
	else:
		continue

	# DB情報取得
	results = sql_run(users, host, dbnames, passwords, factory.factory_code)

	# 材質コードが記載されている行の先頭を取得
	col = factory.zaishitsu_top_cell

	# 材質コードが記載されているセルがある間ループ
	while write_ws.cell(col, 3).value:
		# Excelに記載されている材質コードを取得
		zaishitu_code = int(write_ws.cell(col, 3).value)

		# 取得した材質コードでDBの情報を検索
		found_flg = 0
		for result in results:
			if result[0] == zaishitu_code:
				found_flg = 1
				break

		# DB上に情報があった場合はその情報を、DB上にない場合は0を入力
		if found_flg == 1:
			input_zaishitu = result[0]
			input_jyuryo = result[1]
		else:
			input_zaishitu = zaishitu_code
			input_jyuryo = 0

		# Excelに入力
		# edit_modeが1の場合、DBにない情報はスキップする
		if not (factory.edit_mode == 1 and found_flg == 0):
			input_cell = write_ws.cell(col, row)
			input_cell.number_format = "#,##0"
			input_cell.value = input_jyuryo

		col = col + 1

# Book_write.xlsxに上書保存
try:
	write_wb.save(excel_file)
	print("エクセルにデータを保存しました。")
except(PermissionError):
	print("ファイルの書き込みに失敗しました。")

# エクセルファイルのクローズ
print("エクセルファイルをクローズします。")
write_wb.close()

exit()
