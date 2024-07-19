import psycopg2
import pandas as pd
from openpyxl import load_workbook

# Thay đổi các thông số kết nối theo cấu hình của bạn
db_params = {
    'host': '192.168.160.83',
    'database': 'union_hanbai',
    'user': 'unionplate',
    'password': 'etalpnoinu',
    'port': '5432'
}

# Kết nối đến cơ sở dữ liệu
conn = psycopg2.connect(**db_params)
cursor = conn.cursor()

# Thực hiện truy vấn SQL
query = """select hazai_office_factory_code,count(nyushuko_cf='1' or null) as in,count(nyushuko_cf='2' or null) as out
from offc_trn_hazai_nyushuko
where nyushuko_date= to_char(current_timestamp+cast('-2days' as interval),'yyyy/mm/dd')
--where nyushuko_date = '2023/12/07'
and not hazai_office_factory_code in('590','101') 
group by hazai_office_factory_code
order by hazai_office_factory_code;"""
cursor.execute(query)

# Lấy kết quả của truy vấn
result = cursor.fetchall()

# Tạo DataFrame từ kết quả
df = pd.DataFrame(result, columns=[desc[0] for desc in cursor.description])
# Đọc dữ liệu từ tệp tin Excel đã có
excel_path = r'C:\Users\DSP189\Desktop\工場別端材入出庫数_2023 .xlsx'
existing_df = pd.read_excel(excel_path)



# Lấy dòng đầu tiên và gán vào cột C14 và D14
first_row = df.iloc[0]
""" existing_df.at[13, 'C'] = first_row[0]
existing_df.at[13, 'D'] = first_row[1] """
new_row = pd.Series({'C': first_row[0], 'D': first_row[1]})


""" # Lưu DataFrame vào tệp tin Excel đã có với việc ghi đè sheet cuối cùng
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    writer.book = load_workbook(excel_path)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
     """
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    writer.book = load_workbook(excel_path)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    
    # Ghi DataFrame vào sheet cuối cùng
    existing_df.to_excel(writer, index=False, sheet_name='12月')  # Thay 'Sheet1' bằng tên sheet thực tế của bạn
# Đóng kết nối
cursor.close()
conn.close()
