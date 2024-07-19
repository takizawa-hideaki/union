import sys
import psycopg2
import datetime
from datetime import timedelta
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
current_date = datetime.date(2023, 11, 13)
t = datetime.datetime.now()

# Ví dụ: Tạo vòng lặp và chèn dữ liệu
input_detail =[['corporation_code','office_factory_code','yy_mm_dd','online_create_date','online_create_time','add_juchu_volume','jogai_shiyo_cf','jogai_size_block_no','jogai_flag','create_date','create_time','create_staff']]
row = len(input_detail)
column = len(input_detail[0])       
for i in range(3):
        current_date += timedelta(days=1)
 
        input_detail2 = ['54', '540',(current_date.strftime("%Y/%m/%d")),(t.strftime("%Y/%m/%d")),(t.strftime("%H:%M:%S.%f")[:12]),'5000','6F', '0', '1',(t.strftime("%Y/%m/%d")),(t.strftime("%H:%M:%S")),'トムズイリン']
        for k in range(0,row):
            for j in range(0,column):
                v2=input_detail2[k][j]
                ws.cell(column=j+1, row=k+1, value=v2)

 
    #print("INSERT INTO offc_log_furiwake_kado_rate_up (yy_mm_dd) VALUES (%s);", (current_date.strftime("%Y/%m/%d"),))
    #def greet(current_date):
    ##print ("54, 540,",(current_date.strftime("%Y/%m/%d")),",",(t.strftime("%Y/%m/%d")),(t.strftime("%H:%M:%S.%f")), ",5000,6F, 0, 1,",(t.strftime("%Y/%m/%d")),",",(t.strftime("%H:%M:%S")),",トムズイリン")
   


  #Xác định số hàng và cột lớn nhất trong file excel cần tạo


  #Tạo một workbook mới và active nó

  
  #Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
for i in range(0,row):
    for j in range(0,column):
      v=input_detail[i][j]
      ws.cell(column=j+1, row=i+1, value=v)


  #Lưu lại file Excel
output_excel_path= './user.xlsx'
wb.save(output_excel_path)
##from psycopg2 import extensions

#wb = openpyxl.Workbook()
# Chọn một trang tính (sheet) mặc định
#sheet = wb.active
# PostgreSQL　データベースへの接続を確立する
conn = psycopg2.connect(
    host="#####",
    database="linh_work",
    user="linh",
    password="######"
)

# データベースを操作するためのカーソルを作成する
cur = conn.cursor()

# テーブル
cur.execute(
    "select * from offc_log_furiwake_kado_rate_up where corporation_code = '60' and office_factory_code = '600' /*and jogai_shiyo_cf = '6F'*/  and yy_mm_dd = '2023/11/13' "
)

    # Lấy kết quả của truy vấn
results = cur.fetchall()

    # In kết quả
#for row in results:
 #       print(row)
# Tạo giá trị ngày ban đầu
#time = '2023/11/13'
##current_date = extensions.TimestampFromTicks(time.time())
##current_date = datetime.now()
#current_date = datetime.date(2023, 11, 13).strftime("%Y/%m/%d")

    # Đóng con trỏ và kết nối
cur.close()
conn.close()
#wb.save("example1.xlsx")


exit()

# Tạo giá trị ngày ban đầu
#current_date = datetime.(2023.11.13)

# Số lần lặp và thời gian cách nhau một ngày
num_iterations = 10

# Ví dụ: Tạo vòng lặp và chèn dữ liệu
for i in range(num_iterations):
    cur.execute("INSERT INTO your_table (date_column) VALUES (%s);", (current_date,))
    current_date += timedelta(days=1)

# Lưu thay đổi vào cơ sở dữ liệu
conn.commit()

# Đóng kết nối và con trỏ
cur.close()
conn.close()
