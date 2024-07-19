import psycopg2
from datetime import datetime, timedelta
import openpyxl
import sys

class FactoryInfo:
	def __init__(self, coporation_code, office_factory_code, office_factory_name, office_factory_ryaku, ip_address, nyuko_cell, syuko_cell):
		self.coporation_code = coporation_code
		self.factory_code = office_factory_code
		self.name = office_factory_name
		self.ryaku = office_factory_ryaku
		self.ip = ip_address
		self.nyuko_cell = nyuko_cell
		self.syuko_cell = syuko_cell

# DB接続 → SQL実行 → DB切断
def sql_run(users, host, dbnames, passwords, nyushuko_date):
	# DB接続
	conn = psycopg2.connect(" user=" + users +" host=" + host +" dbname=" + dbnames +" password=" + passwords)

	# sqlの実行
	cur = conn.cursor()
	cur.execute(f"\
				select hazai_office_factory_code,count(nyushuko_cf='1' or null) as in,count(nyushuko_cf='2' or null) as out \
				from offc_trn_hazai_nyushuko \
				where nyushuko_date= '{nyushuko_date}' \
				and not hazai_office_factory_code in('590','101') \
				group by hazai_office_factory_code \
				order by hazai_office_factory_code" 
				)
	results = cur.fetchall()

	# DB切断
	cur.close()
	conn.close()

	return results


# 工場の情報
factorys = [
	#FactoryInfo(コーポレートコード, 工場コード, 工場名, 工場名略, IPアドレス, 入庫セル番号, 出庫セル番号),
	FactoryInfo(10, 100, "㈱ユニオンプレート本社工場", "本社", "192.168.162.51", 3, 4),
	FactoryInfo(10, 101, "ユニオン本社ＦＣ工場", "FC工場", "192.168.162.61", 0, 0),
	FactoryInfo(10, 520, "ユニオン尾道アルミ工場", "尾道アルミ", "192.168.17.61", 13, 14),
	FactoryInfo(10, 570, "ユニオン尾道工場", "尾道", "192.168.17.51", 11, 12),
	FactoryInfo(10, 580, "ユニオン厚木工場", "厚木", "192.168.18.51", 7, 8),
	FactoryInfo(10, 650, "ユニオン小牧工場", "小牧", "192.168.25.51", 9, 10),
	FactoryInfo(10, 670, "ユニオン上田工場", "上田", "192.168.161.51", 5, 6),
	FactoryInfo(10, 680, "ユニオン千曲工場", "千曲", "192.168.28.51", 0, 0),
	FactoryInfo(50, 500, "ＩＮＳ加工工場", "INS", "192.168.10.51", 15, 16),
	FactoryInfo(51, 510, "㈲藤精工", "藤精工", "192.168.11.51", 0, 0),
	FactoryInfo(53, 530, "㈲サンテック", "ｻﾝﾃｯｸ", "192.168.13.51", 17, 18),
	FactoryInfo(54, 540, "㈲エムケイ精工", "MK", "192.168.14.51", 0, 0),
	FactoryInfo(55, 550, "㈱渡辺製作所", "渡辺", "192.168.15.51", 0, 0),
	FactoryInfo(59, 590, "㈲エムテーエス", "MTS", "192.168.19.51", 0, 0),
	FactoryInfo(60, 600, "ＡＵＫプレート㈱", "AUK", "192.168.20.51", 19, 20),
	FactoryInfo(60, 601, "ＡＵＫプレート㈱アルミ工場", "ＡＵＫアルミ", "192.168.20.61", 21, 22),
	FactoryInfo(63, 630, "㈱メカニックメタル", "MM", "192.168.23.51", 23, 24),
	FactoryInfo(64, 640, "㈲鈴木鋼管工業", "鈴木", "192.168.24.51", 0, 0)
]

# 引数の確認（前日以外を指定する場合、引数なしは前日を取得）
if (len(sys.argv) > 1):
	target_day = str(sys.argv[1])
	target_day_1 = datetime.strptime(target_day, "%Y/%m/%d")
	target_month = target_day_1.strftime("%m")
	
	
else:
	# 前日日付の取得とフォーマット変換//*******//
	target_day = datetime.now() - timedelta(days=1)
	target_month = target_day.strftime("%m")
	target_day = target_day.strftime("%Y/%m/%d")
	

# Excelファイルの場所
excel_file = r'\\192.168.160.6\\usbdisk3\\システム運用\\■新システム\\端材管理システム関連\\工場別端材入出庫数_2024.xlsx'

# データベースへの接続用設定
users = 'unionplate'			# ユーザID 
host = '192.168.160.83'			# 接続先IPアドレス
dbnames = 'union_hanbai'		# DB名
passwords = 'etalpnoinu'		# パスワード

# Excelファイルのオープン
# ブックを変数に格納
print("エクセルファイルをオープンします")
try:
	write_wb = openpyxl.load_workbook(excel_file)
except(FileNotFoundError):
	print("書き込むファイルがありません。")
	exit()

# シートを変数に名前で格納


# Excelのデータと比較するため日付フォーマットを変換
target_day_tmp = target_day.replace("'", "")
target_day_tmp = target_day_tmp.replace("/", "-")
write_ws = write_wb[(target_month+"月")]
# 日付をもとに入力する行の取得
found_flg = 0
for col in range(4, 35):
	cell_day = write_ws.cell(col, 2)
	test_cel_day = str(cell_day.value)
	if target_day_tmp in test_cel_day :
		found_flg = 1
		break

# 日付が見つけられなかった時の処理
if found_flg == 0:
	print("入力先が見つかりません。")
	input("j:")

# 入力先の行にデータがあるかどうかチェック（上書きしないように）
for i in range(3,25):
	cell_data = write_ws.cell(col, i)
	if cell_data.value:
		print("既にデータがあります。")
		exit()

# SQLの実行
print("DBよりデータを取得します。")
results = sql_run(users, host, dbnames, passwords, target_day)

# 1レコードもない場合
if len(results) < 1:
	print("取得するデータがありません。")
	exit()

# １レコードエクセルに入力
print("エクセルにデータを入力します。")
for result in results:

	# 取得したDBのレコードと同じ工場の情報を取得
	for factory in factorys:
		if factory.factory_code == result[0]:
			break
	print(factory.ryaku, result)

	# 入庫のセルを取得する
	cell_nyuko_cell = write_ws.cell(col, factory.nyuko_cell)

	# 出庫のセルを取得する
	cell_syuko_cell = write_ws.cell(col, factory.syuko_cell)

	# 入庫セルに値を設定する
	cell_nyuko_cell.value = result[1]

	# 出庫セルに値を設定する
	cell_syuko_cell.value = result[2]

# Book_write.xlsxに上書保存
try:
	write_wb.save(excel_file)
	print("エクセルにデータを保存しました。")
except(PermissionError):
	print("ファイルの書き込みに失敗しました。")
	input("j:")

# エクセルファイルのクローズ
print("エクセルファイルをクローズします。")
write_wb.close()
